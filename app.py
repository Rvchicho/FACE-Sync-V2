import io
import os
import openpyxl
import pandas as pd
from PIL import Image
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

from lector_mayor import cargar_mayor_presupuestario, limpiar_monto_boliviano

st.set_page_config(page_title="FACE-Sync V2 - MVP PNUD", page_icon="📊", layout="wide")
# CSS exclusivo para ajustar el tamaño del texto de las tarjetas de métricas
st.markdown("""
    <style>
        [data-testid="stMetricValue"] {
            font-size: 1.6rem !important;
            font-weight: 700 !important;
        }
        [data-testid="stMetricLabel"] {
            font-size: 0.85rem !important;
        }
    </style>
""", unsafe_allow_html=True)

# Búsqueda flexible del Logo
logo_img = None
rutas_logo = [
    "Logo_PROCOSI_OK_1.png",
    os.path.join("assets", "Logo_PROCOSI_OK_1.png"),
    os.path.join("assets", "logo.png"),
    "input_file_0.png"
]

for r in rutas_logo:
    if os.path.exists(r):
        try:
            logo_img = Image.open(r)
            break
        except Exception:
            pass

# --- BARRA LATERAL ---
with st.sidebar:
    if logo_img:
        st.image(logo_img, use_container_width=100)
    
    st.title("⚙️ Panel de Control")
    st.markdown("---")
    file_mayor = st.file_uploader("1. Mayor Presupuestario (VISUAL)", type=["xlsx", "xls"], accept_multiple_files=False)
    file_plantilla = st.file_uploader("2. Plantilla Oficial FACE (Excel)", type=["xlsx"], accept_multiple_files=False)
    
    with st.expander("🔄 Actualizar Matriz Budget (Opcional)"):
        file_budget = st.file_uploader("Subir nuevo Budget.csv solo si el proyecto cambió", type=["csv"], accept_multiple_files=False)
        st.caption("Por defecto, se utiliza el catálogo interno almacenado en data/.")
    
    st.markdown("---")
    placeholder_boton = st.empty()
    
    with st.expander("💡 Formato Recomendado de Glosa"):
        st.write("""
        `[Proveedor] / [N° Factura] / [Descripción] / [N° Cheque] / [Activity ID]`
        
        *Etiquetas opcionales:*
        - `ACT:` Activity ID
        - `F:` Factura
        - `CH:` Cheque
        - `P:` Proveedor
        """)


def generar_excel_face(plantilla_bytes, df_datos):
    wb = openpyxl.load_workbook(io.BytesIO(plantilla_bytes))
    nombre_pestana = "SR_Transaction_Details" if "SR_Transaction_Details" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[nombre_pestana]
    fila_inicio = 22
    
    for i, row in df_datos.iterrows():
        r = fila_inicio + i
        ws.cell(row=r, column=2, value=i + 1)
        
        cell_date = ws.cell(row=r, column=3, value=row["Fecha"])
        cell_date.number_format = 'yyyy-mm-dd'
        
        ws.cell(row=r, column=4, value=str(row["Voucher No"]))
        ws.cell(row=r, column=5, value=str(row["Payment / Ref"]))
        ws.cell(row=r, column=6, value=str(row["Payee / Vendor"]))
        ws.cell(row=r, column=7, value=str(row["Description"]))
        
        b_line = int(row["Budget Line No"]) if str(row["Budget Line No"]).isdigit() else row["Budget Line No"]
        ws.cell(row=r, column=8, value=b_line)
        ws.cell(row=r, column=9, value=str(row["Account"]))
        ws.cell(row=r, column=14, value=float(row["Payment (Bs)"]))
        ws.cell(row=r, column=17, value=float(row["Expenses (87%)"]))
        ws.cell(row=r, column=18, value=float(row["VAT Tax (13%)"]))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def obtener_presupuesto_total_k22(file_plantilla_bytes, df_budget_cat):
    """
    Mantiene un techo de presupuesto oficial y constante (1.564.000 Bs) 
    o lo lee directamente de la celda K22 si la plantilla Excel está cargada.
    """
    if file_plantilla_bytes:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_plantilla_bytes), data_only=True)
            nombre_pestana = "SR_Transaction_Details" if "SR_Transaction_Details" in wb.sheetnames else wb.sheetnames[0]
            ws = wb[nombre_pestana]
            
            # Celda K22 (Columna K = 11, Fila = 22)
            val_k22 = ws.cell(row=22, column=11).value
            monto_k22 = limpiar_monto_boliviano(val_k22)
            if monto_k22 > 0:
                return monto_k22
        except Exception:
            pass
            
    # Si aún no han subido la plantilla FACE, fijar el Techo Oficial PNUD (1,564,455.24 Bs)
    # para que la cifra no 'salte' al subir la plantilla
    return 1564455.24


# --- ENCABEZADO PRINCIPAL ---
col_head1, col_head2 = st.columns([1, 4])
with col_head1:
    if logo_img:
        st.image(logo_img, width=120)
with col_head2:
    st.title("FACE-Sync V2 (PNUD - PROCOSI)")
    st.caption("Conector Financiero para la Automatización de Reportes FACE (PNUD)")

if file_mayor:
    try:
        df_procesado, df_budget_cat = cargar_mayor_presupuestario(file_mayor, file_budget if 'file_budget' in locals() and file_budget else None)
        
        if not df_procesado.empty:
            plantilla_bytes = file_plantilla.read() if file_plantilla else None
            
            if plantilla_bytes:
                excel_generado = generar_excel_face(plantilla_bytes, df_procesado)
                placeholder_boton.download_button(
                    label="🚀 DESCARGAR FACE LLENADO (.XLSX)",
                    data=excel_generado,
                    file_name="FACE_PNUD_Llenado.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary"
                )
            else:
                placeholder_boton.warning("⚠️ Carga la Plantilla FACE para descargar.")

            # Lectura del techo de presupuesto desde la celda K22 de la plantilla (o respaldo en CSV)
            presupuesto_total = obtener_presupuesto_total_k22(plantilla_bytes, df_budget_cat)
            ejecutado_total = df_procesado['Payment (Bs)'].sum()
            porcentaje_avance = (ejecutado_total / presupuesto_total * 100) if presupuesto_total > 0 else 0.0

            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Presupuesto Total Proyecto ", f"Bs {presupuesto_total:,.2f}" if presupuesto_total > 0 else "N/D")
            m2.metric("Ejecutado Período", f"Bs {ejecutado_total:,.2f}")
            m3.metric("% Absorción de Fondos", f"{porcentaje_avance:.1f}%" if presupuesto_total > 0 else "N/D")
            m4.metric("Total Transacciones", len(df_procesado))
            
            st.markdown("---")
            
            tab_visor, tab_dash = st.tabs(["📋 Visor de Transacciones", "📈 Dashboard de Avance por Activity ID"])
            
            with tab_visor:
                st.dataframe(
                    df_procesado[[
                        "Voucher No", "Fecha", "Account", "Activity ID", 
                        "Budget Line No", "Payee / Vendor", "Invoice No", 
                        "Description", "Payment / Ref", "Payment (Bs)", 
                        "Expenses (87%)", "VAT Tax (13%)"
                    ]], 
                    use_container_width=True,
                    height=500
                )

            with tab_dash:
                st.subheader("Análisis Individual por Activity ID")
                
                actividades_disponibles = sorted(df_procesado["Activity ID"].unique(), key=lambda x: int(x) if str(x).isdigit() else 0)
                selected_act = st.selectbox("Selecciona un Activity ID para auditar:", actividades_disponibles)
                
                df_act_trans = df_procesado[df_procesado["Activity ID"] == selected_act]
                monto_ejecutado_act = df_act_trans["Payment (Bs)"].sum()
                
                monto_presupuestado_act = 0.0
                info_actividad = {}
                
                if not df_budget_cat.empty:
                    df_budget_act = df_budget_cat[df_budget_cat["Activity ID"].astype(str).str.strip().str.split('.').str[0] == str(selected_act)]
                    if not df_budget_act.empty:
                        monto_presupuestado_act = df_budget_act[' Amount Total '].apply(limpiar_monto_boliviano).sum()
                        first_row = df_budget_act.iloc[0]
                        info_actividad = {
                            "Module": first_row.get("Module", "N/A"),
                            "Intervention": first_row.get("Intervention", "N/A"),
                            "Activity": first_row.get("Activity", "N/A"),
                            "Cost input": first_row.get("Cost input", "N/A")
                        }

                st.info(f"""
                **Detalle de la Actividad (Activity ID {selected_act}):**
                * **Módulo:** {info_actividad.get('Module', 'N/A')}
                * **Intervención:** {info_actividad.get('Intervention', 'N/A')}
                * **Actividad:** {info_actividad.get('Activity', 'N/A')}
                * **Cost Input:** {info_actividad.get('Cost input', 'N/A')}
                """)
                
                c_kpi1, c_kpi2, c_kpi3 = st.columns(3)
                c_kpi1.metric("Presupuesto de la Actividad", f"Bs {monto_presupuestado_act:,.2f}" if monto_presupuestado_act > 0 else "N/D")
                c_kpi2.metric("Monto Ejecutado", f"Bs {monto_ejecutado_act:,.2f}")
                saldo_act = monto_presupuestado_act - monto_ejecutado_act 
                c_kpi3.metric("Saldo Disponible", f"Bs {saldo_act:,.2f}" if monto_presupuestado_act > 0 else "N/D")
                
                st.markdown("---")
                
                # --- EDICIÓN DE LA GRÁFICA COMPARATIVA ---
                if monto_presupuestado_act > 0 or monto_ejecutado_act > 0:
                    fig_comp = go.Figure(data=[
                        go.Bar(
                            name='Presupuesto Aprobado', 
                            x=[f'Actividad {selected_act}'], 
                            y=[monto_presupuestado_act], 
                            marker_color='#1E3A8A'  # <-- CAMBIA AQUÍ EL COLOR (Azul Oscuro)
                        ),
                        go.Bar(
                            name='Monto Ejecutado', 
                            x=[f'Actividad {selected_act}'], 
                            y=[monto_ejecutado_act], 
                            marker_color='#D97706'  # <-- CAMBIA AQUÍ EL COLOR (Naranja / Ámbar)
                        )
                    ])
                    fig_comp.update_layout(
                        barmode='group',
                        title=f'Comparativa: Presupuesto vs. Ejecución (Activity ID {selected_act})',
                        yaxis_title='Monto (Bs)',
                        height=300,        # <-- CAMBIA AQUÍ LA ALTURA (en píxeles)
                        bargap=0.4,        # <-- CAMBIA AQUÍ EL ESPACIO/ANCHO DE LAS BARRAS (0.1 a 0.8)
                        margin=dict(l=20, r=20, t=40, b=20) # Ajuste de márgenes
                    )
                    
                    # Para controlar el ancho en pantalla:
                    col_chart, _ = st.columns([2, 1]) # 2/3 del ancho disponible
                    with col_chart:
                        st.plotly_chart(fig_comp, use_container_width=True)
                
                st.write("#### Transacciones Asociadas a esta Actividad")
                st.dataframe(df_act_trans[["Voucher No", "Account", "Budget Line No", "Payee / Vendor", "Description", "Payment / Ref", "Payment (Bs)"]], use_container_width=True)

    except Exception as e:
        st.error(f"Error al procesar el archivo: {str(e)}")
else:
    st.info("👋 Por favor, sube el archivo Excel del Mayor Presupuestario en el menú lateral para comenzar.")
